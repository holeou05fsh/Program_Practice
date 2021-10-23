def Dice_Game():
    import random
    # a=b=c=d = random.randint(1,6)
    a = random.randint(1, 6)
    b = random.randint(1, 6)
    c = random.randint(1, 6)
    d = random.randint(1, 6)

    dice_1 = [a, b, c, d]
    dice_1a = set(dice_1)
    dice_1f = []

    print(dice_1)
    for each1_a in dice_1a:
        count = 0
        for each_1b in dice_1:
            if each1_a == each_1b:
                count += 1
                if count == 4:
                    dice_1f.append('一色')
                    dice_1f.append(each1_a)
                elif count == 3:
                    dice_1f.append('三色')
                elif count == 2:
                    dice_1f.append('一對')
                    dice_1f.append(each1_a)
                elif count == 1:
                    dice_1f.append(each1_a)

    if dice_1f.count('一色') == 1:
        dice_result = ['一色:' + str(dice_1[1])]
        print('一色:', dice_1[1])
    elif dice_1f.count('三色') == 1:
        dice_result = ['在骰一次']
        print('在骰一次')
    elif dice_1f.count('一對') == 2:
        r = dice_1f.index('一對') + 1
        s = dice_1f.index('一對', r) + 1
        if dice_1f[r] > dice_1f[s]:
            dice_result = [dice_1f[1] * 2]
            print('數字:', dice_1f[1] * 2)
        else:
            dice_result = [dice_1f[3] * 2]
            print('數字:', dice_1f[3] * 2)
    elif dice_1f.count('一對') == 1:
        r = int(dice_1f.index('一對')) + 1
        r = dice_1f[r]
        s = dice_1f.index('一對')
        # print(dice_1f)
        del dice_1f[s]
        dice_1f.remove(r)
        dice_1f.remove(r)
        # print(dice_1f)
        dice_1f1 = dice_1f[0]
        dice_1f1 = int(dice_1f[0])
        dice_1f2 = dice_1f[1]
        dice_1f2 = int(dice_1f[1])
        dice_result = [dice_1f1 + dice_1f2]
        print('數字:', dice_1f1 + dice_1f2)
    else:
        dice_result = ['在骰一次']
        print(dice_result[0])

    return dice_result[0]

def player():
    dice_result2 = []
    start = input('開始(Y / N):')
    start = start.upper()
    if start == 'Y':
        dice_result = Dice_Game()
        dice_result2 = str(dice_result)
        dice_result
        # print(agn)
        for i in range(20):
            if dice_result == '在骰一次':
                agn_dice = input('繼續(Y / N):')
                agn_dice = agn_dice.upper()
                if agn_dice == 'Y':
                    dice_result = Dice_Game()
                    dice_result2 = str(dice_result)
                    # print(dice_result)
                    dice_result
                else:
                    dice_result2 = str(0)
                    break
            else:
                break
    else:
        dice_result2 = str(0)
    return dice_result2


Player = input('請輸入玩家有幾位:')
player_statistics = []
NOP = 1
for i in range(int(Player)):
    print('')
    print('玩家', NOP)
    player_result = player()
    player_statistics.append(player_result)
    player_result
    NOP += 1
# print(player_result)
NOP = 0
print('')
# print(player_statistics)
for i in range(int(Player)):
    print('玩家', NOP + 1, '成績:', player_statistics[NOP])
    NOP += 1

a = 1
for i in player_statistics:
    if '一色:' + str(a) in player_statistics:
        b = True
        break
    a += 1
    b = False

if b:
    a = 0
    W_P = [i for i, v in enumerate(player_statistics) if v[0] == '一']
    for i in W_P:
        print('※恭喜 玩家' + str(int(W_P[a]) + 1), '獲勝')
        a += 1
else:
    player_statistics = list(map(int, player_statistics))

    a = 0
    W_P = [i for i, v in enumerate(player_statistics) if v == max(player_statistics)]
    for i in W_P:
        print('※恭喜 玩家' + str(int(W_P[a]) + 1), '獲勝!!!')
        a += 1


import openpyxl
import os
import time

if not os.path.isfile('十八豆紀錄.xlsx'):
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='紀錄')
    ws = wb.active
    ws.append(['日期', '時間', '玩家1', '玩家2', '玩家3', '玩家4', '玩家5'])
    wb.save('十八豆紀錄.xlsx')

wb = openpyxl.load_workbook('十八豆紀錄.xlsx')
sheet = wb['紀錄']
r = str(sheet.max_row + 1)
# print(r,type(r))
ws = wb.active

t = time.localtime()
ws['A' + r] = time.strftime('%Y/%m/%d', t)
ws['B' + r] = time.strftime('%H:%M', t)
ws['C' + r] = time.strftime(str(player_statistics[0]))
ws['D' + r] = time.strftime(str(player_statistics[1]))
ws['E' + r] = time.strftime(str(player_statistics[2]))
ws['F' + r] = time.strftime(str(player_statistics[3]))
ws['G' + r] = time.strftime(str(player_statistics[4]))
wb.save('十八豆紀錄.xlsx')